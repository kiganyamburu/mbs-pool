import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime
import matplotlib.dates as mdates

# Read the projected cashflow Excel file
wb = openpyxl.load_workbook("Sample MBS Projected Cashflow.xlsx", data_only=True)

# Define colors and labels for each scenario
scenarios = {
    "-300 MED ": {"color": "#e41a1c", "label": "-300 MED (WAL: 2.53 yrs)"},
    "- 200 MED": {"color": "#377eb8", "label": "-200 MED (WAL: 4.25 yrs)"},
    "- 100 MED ": {"color": "#4daf4a", "label": "-100 MED (WAL: 8.15 yrs)"},
    "0 MED ": {"color": "#984ea3", "label": "0 MED (WAL: 9.89 yrs)"},
    "+ 100 MED ": {"color": "#ff7f00", "label": "+100 MED (WAL: 10.90 yrs)"},
    "+ 200 MED": {"color": "#ffff33", "label": "+200 MED (WAL: 11.58 yrs)"},
    "+ 300 MED": {"color": "#a65628", "label": "+300 MED (WAL: 12.03 yrs)"},
}

# Create figure
fig, ax = plt.subplots(figsize=(14, 8))

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))

    dates = []
    cashflows = []

    for row in all_rows[2:]:  # Skip header and totals
        if row[0] and row[0] != "Dates":
            try:
                date_str = str(row[0])
                if "/" in date_str:
                    payment_date = datetime.strptime(date_str, "%m/%d/%Y")
                    cashflow = row[5] if row[5] else 0  # Cashflow column

                    dates.append(payment_date)
                    cashflows.append(cashflow)
            except Exception as e:
                continue

    # Get scenario info
    info = scenarios.get(sheet_name, {"color": "gray", "label": sheet_name})

    # Plot only first 120 months (10 years) for better visibility
    plot_dates = dates[:120]
    plot_cashflows = cashflows[:120]

    ax.plot(
        plot_dates,
        plot_cashflows,
        label=info["label"],
        color=info["color"],
        linewidth=1.5,
    )

ax.set_xlabel("Date", fontsize=12)
ax.set_ylabel("Monthly Cashflow ($)", fontsize=12)
ax.set_title(
    "MBS Pool Projected Cashflow by Interest Rate Scenario\n(CUSIP 3140A02B4, Coupon 5.50%, First 10 Years)",
    fontsize=14,
    fontweight="bold",
)
ax.legend(loc="upper right", fontsize=9)
ax.grid(True, alpha=0.3)

# Format x-axis
ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
ax.xaxis.set_major_locator(mdates.YearLocator())
plt.xticks(rotation=45)

# Format y-axis with comma separators
ax.get_yaxis().set_major_formatter(plt.FuncFormatter(lambda x, p: format(int(x), ",")))

plt.tight_layout()
plt.savefig("Q3-3_Cashflow_Chart.png", dpi=150, bbox_inches="tight")
print("Chart saved as Q3-3_Cashflow_Chart.png")

# Also create a summary statistics output
print("\nCashflow Summary Statistics:")
print("=" * 60)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    totals_row = all_rows[1]  # Second row contains totals
    print(
        f"{sheet_name.strip()}: Total Cashflow = ${totals_row[5]:,.0f}"
        if totals_row[5]
        else f"{sheet_name.strip()}: N/A"
    )
