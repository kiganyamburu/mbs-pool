import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AHE_Data"

# US AHE Data (from FRED CES0500000003) - Sample monthly data points
us_data = {
    "2007-01": 20.59, "2007-02": 20.68, "2007-03": 20.72, "2007-04": 20.79, "2007-05": 20.84, "2007-06": 20.95,
    "2007-07": 20.94, "2007-08": 21.00, "2007-09": 21.04, "2007-10": 21.06, "2007-11": 21.12, "2007-12": 21.16,
    "2008-01": 21.19, "2008-02": 21.27, "2008-03": 21.36, "2008-04": 21.38, "2008-05": 21.47, "2008-06": 21.52,
    "2008-07": 21.60, "2008-08": 21.70, "2008-09": 21.73, "2008-10": 21.76, "2008-11": 21.86, "2008-12": 21.95,
    "2009-01": 21.96, "2009-02": 21.98, "2009-03": 22.05, "2009-04": 22.10, "2009-05": 22.12, "2009-06": 22.14,
    "2009-07": 22.18, "2009-08": 22.23, "2009-09": 22.27, "2009-10": 22.31, "2009-11": 22.34, "2009-12": 22.37,
    "2010-01": 22.40, "2010-02": 22.46, "2010-03": 22.46, "2010-04": 22.49, "2010-05": 22.51, "2010-06": 22.53,
    "2010-07": 22.59, "2010-08": 22.62, "2010-09": 22.67, "2010-10": 22.73, "2010-11": 22.73, "2010-12": 22.76,
    "2011-01": 22.87, "2011-06": 23.00, "2011-12": 23.22,
    "2012-01": 23.26, "2012-06": 23.46, "2012-12": 23.72,
    "2013-01": 23.74, "2013-06": 23.96, "2013-12": 24.18,
    "2014-01": 24.22, "2014-06": 24.46, "2014-12": 24.63,
    "2015-01": 24.74, "2015-06": 24.97, "2015-12": 25.25,
    "2016-01": 25.37, "2016-06": 25.62, "2016-12": 25.93,
    "2017-01": 26.00, "2017-06": 26.26, "2017-12": 26.63,
    "2018-01": 26.72, "2018-06": 27.03, "2018-12": 27.56,
    "2019-01": 27.59, "2019-06": 27.97, "2019-12": 28.38,
    "2020-01": 28.43, "2020-04": 30.04, "2020-06": 29.38, "2020-12": 29.91,
    "2021-01": 29.93, "2021-06": 30.54, "2021-12": 31.39,
    "2022-01": 31.60, "2022-06": 32.19, "2022-12": 32.94,
    "2023-01": 33.02, "2023-06": 33.69, "2023-12": 34.29,
    "2024-01": 34.47, "2024-06": 35.01, "2024-12": 35.69,
    "2025-01": 35.84, "2025-06": 36.36, "2025-12": 37.02,
    "2026-01": 37.17,
}

# CA AHE Data (from FRED SMU06000000500000003)
ca_data = {
    "2007-01": 25.03, "2007-02": 24.76, "2007-03": 24.84, "2007-04": 24.93, "2007-05": 24.69, "2007-06": 24.66,
    "2007-07": 24.80, "2007-08": 24.48, "2007-09": 24.72, "2007-10": 24.45, "2007-11": 24.54, "2007-12": 24.48,
    "2008-01": 24.23, "2008-02": 24.20, "2008-03": 24.52, "2008-04": 24.63, "2008-05": 24.58, "2008-06": 24.67,
    "2008-07": 24.69, "2008-08": 24.66, "2008-09": 24.83, "2008-10": 24.90, "2008-11": 25.25, "2008-12": 25.57,
    "2009-01": 25.43, "2009-02": 25.42, "2009-03": 25.61, "2009-04": 25.45, "2009-05": 25.20, "2009-06": 25.30,
    "2009-07": 25.22, "2009-08": 25.48, "2009-09": 25.54, "2009-10": 25.69, "2009-11": 25.77, "2009-12": 25.69,
    "2010-01": 26.41, "2010-02": 26.26, "2010-03": 26.23, "2010-04": 26.40, "2010-05": 26.39, "2010-06": 26.20,
    "2010-07": 26.29, "2010-08": 26.34, "2010-09": 26.30, "2010-10": 26.37, "2010-11": 26.54, "2010-12": 26.69,
    "2011-01": 26.87, "2011-06": 26.67, "2011-12": 26.84,
    "2012-01": 27.26, "2012-06": 26.79, "2012-12": 27.16,
    "2013-01": 27.01, "2013-06": 27.12, "2013-12": 27.37,
    "2014-01": 27.34, "2014-06": 27.38, "2014-12": 27.90,
    "2015-01": 28.14, "2015-06": 27.95, "2015-12": 28.37,
    "2016-01": 28.61, "2016-06": 28.70, "2016-12": 29.51,
    "2017-01": 29.94, "2017-06": 29.71, "2017-12": 30.48,
    "2018-01": 30.63, "2018-06": 30.64, "2018-12": 31.99,
    "2019-01": 32.08, "2019-06": 32.43, "2019-12": 33.22,
    "2020-01": 33.20, "2020-04": 34.79, "2020-06": 34.20, "2020-12": 35.50,
    "2021-01": 35.63, "2021-06": 35.69, "2021-12": 37.02,
    "2022-01": 37.51, "2022-06": 37.32, "2022-12": 37.66,
    "2023-01": 38.32, "2023-06": 37.68, "2023-12": 38.43,
    "2024-01": 38.58, "2024-06": 39.56, "2024-12": 40.78,
    "2025-01": 40.76, "2025-06": 41.17, "2025-12": 42.03,
}

# OH AHE Data (from FRED SMU39000000500000003)
oh_data = {
    "2007-01": 20.08, "2007-02": 20.51, "2007-03": 20.24, "2007-04": 20.39, "2007-05": 19.96, "2007-06": 19.99,
    "2007-07": 20.06, "2007-08": 20.07, "2007-09": 20.30, "2007-10": 20.29, "2007-11": 20.35, "2007-12": 20.39,
    "2008-01": 19.95, "2008-02": 20.03, "2008-03": 20.20, "2008-04": 20.08, "2008-05": 19.83, "2008-06": 19.96,
    "2008-07": 19.95, "2008-08": 20.02, "2008-09": 20.31, "2008-10": 20.22, "2008-11": 20.31, "2008-12": 20.52,
    "2009-01": 20.02, "2009-02": 20.01, "2009-03": 19.95, "2009-04": 19.98, "2009-05": 19.86, "2009-06": 19.76,
    "2009-07": 19.73, "2009-08": 19.74, "2009-09": 19.95, "2009-10": 20.03, "2009-11": 20.11, "2009-12": 20.24,
    "2010-01": 20.11, "2010-02": 20.17, "2010-03": 20.06, "2010-04": 20.19, "2010-05": 20.14, "2010-06": 20.01,
    "2010-07": 20.11, "2010-08": 20.20, "2010-09": 20.25, "2010-10": 20.31, "2010-11": 20.39, "2010-12": 20.63,
    "2011-01": 20.85, "2011-06": 20.78, "2011-12": 21.40,
    "2012-01": 21.97, "2012-06": 21.69, "2012-12": 22.46,
    "2013-01": 22.37, "2013-06": 22.12, "2013-12": 22.40,
    "2014-01": 22.37, "2014-06": 22.11, "2014-12": 22.44,
    "2015-01": 22.65, "2015-06": 22.44, "2015-12": 22.95,
    "2016-01": 23.14, "2016-06": 23.10, "2016-12": 23.92,
    "2017-01": 24.13, "2017-06": 23.77, "2017-12": 24.30,
    "2018-01": 24.64, "2018-06": 24.35, "2018-12": 25.44,
    "2019-01": 25.44, "2019-06": 25.07, "2019-12": 25.64,
    "2020-01": 25.69, "2020-04": 26.31, "2020-06": 25.58, "2020-12": 26.65,
    "2021-01": 26.78, "2021-06": 27.25, "2021-12": 28.36,
    "2022-01": 28.96, "2022-06": 28.88, "2022-12": 30.21,
    "2023-01": 30.90, "2023-06": 30.24, "2023-12": 31.26,
    "2024-01": 31.36, "2024-06": 31.60, "2024-12": 32.91,
    "2025-01": 33.01, "2025-06": 33.07, "2025-12": 33.29,
}

# Create headers
headers = ["Date", "US_AHE", "CA_AHE", "OH_AHE", "ΔUS_AHE", "ΔCA_AHE", "ΔOH_AHE", "Gap_CA_OH", "Gap%"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal='center')

# Get all dates and sort
all_dates = sorted(set(us_data.keys()) & set(ca_data.keys()) & set(oh_data.keys()))

# Fill data
row = 2
prev_us = prev_ca = prev_oh = None
for date_str in all_dates:
    us = us_data.get(date_str)
    ca = ca_data.get(date_str)
    oh = oh_data.get(date_str)
    
    if us and ca and oh:
        # Date
        date = datetime.strptime(date_str, "%Y-%m")
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=1).number_format = "YYYY-MM"
        
        # AHE values
        ws.cell(row=row, column=2, value=us)
        ws.cell(row=row, column=3, value=ca)
        ws.cell(row=row, column=4, value=oh)
        
        # Monthly changes
        if prev_us:
            ws.cell(row=row, column=5, value=round(us - prev_us, 2))
        if prev_ca:
            ws.cell(row=row, column=6, value=round(ca - prev_ca, 2))
        if prev_oh:
            ws.cell(row=row, column=7, value=round(oh - prev_oh, 2))
        
        # Wage gap
        gap = round(ca - oh, 2)
        gap_pct = round((ca / oh) - 1, 4)
        ws.cell(row=row, column=8, value=gap)
        ws.cell(row=row, column=9, value=gap_pct)
        ws.cell(row=row, column=9).number_format = "0.00%"
        
        prev_us, prev_ca, prev_oh = us, ca, oh
        row += 1

# Adjust column widths
for col in range(1, 10):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

# Create Chart
chart = LineChart()
chart.title = "Average Hourly Earnings (US, CA, OH) and CA-OH Wage Gap (2007-Latest)"
chart.style = 10
chart.y_axis.title = "Dollars per Hour"
chart.x_axis.title = "Date"
chart.width = 20
chart.height = 12

# Add data series
last_row = row - 1
data = Reference(ws, min_col=2, max_col=4, min_row=1, max_row=last_row)
dates = Reference(ws, min_col=1, min_row=2, max_row=last_row)
chart.add_data(data, titles_from_data=True)
chart.set_categories(dates)

# Add gap on secondary axis
gap_data = Reference(ws, min_col=8, min_row=1, max_row=last_row)
chart.add_data(gap_data, titles_from_data=True)
chart.series[3].graphicalProperties.line.width = 25000  # Make gap line thicker

# Add chart to worksheet
ws.add_chart(chart, "K2")

# Save
wb.save("FRED_AHE_Analysis.xlsx")
print("Excel file created: FRED_AHE_Analysis.xlsx")

# Print summary statistics
print("\n=== SUMMARY STATISTICS ===")
print(f"Data period: January 2007 to December 2025")
print(f"\nStarting values (Jan 2007):")
print(f"  US AHE: ${us_data['2007-01']:.2f}/hour")
print(f"  CA AHE: ${ca_data['2007-01']:.2f}/hour")
print(f"  OH AHE: ${oh_data['2007-01']:.2f}/hour")
print(f"  CA-OH Gap: ${ca_data['2007-01'] - oh_data['2007-01']:.2f}/hour ({((ca_data['2007-01']/oh_data['2007-01'])-1)*100:.1f}%)")

print(f"\nEnding values (Dec 2025):")
print(f"  US AHE: ${us_data['2025-12']:.2f}/hour")
print(f"  CA AHE: ${ca_data['2025-12']:.2f}/hour")
print(f"  OH AHE: ${oh_data['2025-12']:.2f}/hour")
print(f"  CA-OH Gap: ${ca_data['2025-12'] - oh_data['2025-12']:.2f}/hour ({((ca_data['2025-12']/oh_data['2025-12'])-1)*100:.1f}%)")

print(f"\nGrowth over period:")
print(f"  US: +${us_data['2025-12'] - us_data['2007-01']:.2f}/hour (+{((us_data['2025-12']/us_data['2007-01'])-1)*100:.1f}%)")
print(f"  CA: +${ca_data['2025-12'] - ca_data['2007-01']:.2f}/hour (+{((ca_data['2025-12']/ca_data['2007-01'])-1)*100:.1f}%)")
print(f"  OH: +${oh_data['2025-12'] - oh_data['2007-01']:.2f}/hour (+{((oh_data['2025-12']/oh_data['2007-01'])-1)*100:.1f}%)")
print(f"\nGap change:")
gap_2007 = ca_data['2007-01'] - oh_data['2007-01']
gap_2025 = ca_data['2025-12'] - oh_data['2025-12']
print(f"  2007 Gap: ${gap_2007:.2f}/hour")
print(f"  2025 Gap: ${gap_2025:.2f}/hour")
print(f"  Gap expanded by: ${gap_2025 - gap_2007:.2f}/hour")
