import openpyxl

wb = openpyxl.load_workbook(
    "Sample MBS actual paydown FN_BJ97911_MTGE PDI .xlsx", data_only=True
)

# Get Tab #2 data (MBS Pool PAYDOWN history)
ws2 = wb["#2 MBS Pool PAYDOWN history "]
paydown_data = []
for row in list(ws2.iter_rows(values_only=True))[3:]:
    if row[0] and row[5]:
        paydown_data.append({"date": str(row[0])[:10], "balance": row[5]})

# Get Tab #3 data (MBS Collateral history)
ws3 = wb["#3 MBS Collateral history"]
collateral_data = []
for row in list(ws3.iter_rows(values_only=True))[3:]:
    if row[0] and row[6]:
        collateral_data.append({"date": str(row[0]), "balance": row[6]})

print("Comparison of MBS Pool Balance (Tab #2) vs Collateral Balance (Tab #3):")
print("=" * 70)

for i in range(min(15, len(paydown_data), len(collateral_data))):
    p = paydown_data[i]
    c = collateral_data[i]
    diff = abs(float(p["balance"]) - float(c["balance"]))
    same = "SAME" if diff < 0.01 else "DIFF"
    print(
        f"{c['date']:<12} Pool: {p['balance']:<15} Collateral: {c['balance']:<15} {same}"
    )
